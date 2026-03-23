"""Excel rapor uretici."""

import logging
import os
from datetime import datetime
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.profiler import ColumnProfile, DatabaseProfile, SchemaProfile, TableProfile

logger = logging.getLogger(__name__)

# Stiller
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
PK_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
FK_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
GRADE_FILLS = {
    "A": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    "B": PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"),
    "C": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
    "D": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
    "F": PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid"),
    "N/A": PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid"),
}
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")


class ExcelReportGenerator:
    """openpyxl tabanli Excel rapor uretici."""

    def __init__(self, mapping_enabled: bool = False):
        self.mapping_enabled = mapping_enabled

    def generate(self, profile: DatabaseProfile, output_path: str) -> str:
        """Excel raporu uret. Dosya yolunu dondur."""
        wb = Workbook()

        # Sheet 1: Ozet
        self._write_summary(wb, profile)

        # Sheet 2: Schema Ozet
        self._write_schema_summary(wb, profile)

        # Sheet 3: Tablo Profil
        self._write_table_profile(wb, profile)

        # Sheet 4: Kolon Profil
        self._write_column_profile(wb, profile)

        # Sheet 5: Top Degerler
        self._write_top_values(wb, profile)

        # Sheet 6: Pattern Analiz
        self._write_pattern_analysis(wb, profile)

        # Sheet 7: Outlier Rapor
        self._write_outlier_report(wb, profile)

        # Varsayilan bos sheet'i sil
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
        wb.save(output_path)
        logger.info("Excel rapor olusturuldu: %s", output_path)
        return output_path

    def _apply_header(self, ws, headers: List[str], row: int = 1) -> None:
        """Baslik satirini formatla."""
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.freeze_panes = ws.cell(row=row + 1, column=1)
        ws.auto_filter.ref = ws.dimensions

    def _auto_width(self, ws, max_width: int = 40) -> None:
        """Kolon genisliklerini otomatik ayarla."""
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 3, max_width)

    def _write_summary(self, wb: Workbook, profile: DatabaseProfile) -> None:
        """Ozet sheeti."""
        ws = wb.create_sheet("Ozet", 0)
        data = [
            ["Proje", "DWH Kaynak Profilleme"],
            ["Veritabani", f"{profile.db_alias} ({profile.db_name})"],
            ["Host", profile.host],
            ["Profilleme Tarihi", profile.profiled_at],
            ["Toplam Sema", profile.total_schemas],
            ["Toplam Tablo", profile.total_tables],
            ["Toplam Kolon", profile.total_columns],
            ["Toplam Satir", f"{profile.total_rows:,}"],
            ["Toplam Boyut", profile.total_size_display or "-"],
            ["Genel Kalite Skoru", f"{profile.overall_quality_score:.2%}"],
        ]
        for row_idx, (label, value) in enumerate(data, 1):
            ws.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row_idx, column=2, value=value)

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 40

    def _write_schema_summary(self, wb: Workbook, profile: DatabaseProfile) -> None:
        """Schema Ozet sheeti."""
        ws = wb.create_sheet("Schema Ozet")
        headers = ["Sema", "Tablo Sayisi", "Toplam Satir", "Toplam Boyut", "Kalite Skoru", "Kalite Notu"]
        self._apply_header(ws, headers)

        for row_idx, schema in enumerate(profile.schemas, 2):
            ws.cell(row=row_idx, column=1, value=schema.schema_name)
            ws.cell(row=row_idx, column=2, value=schema.table_count)
            ws.cell(row=row_idx, column=3, value=schema.total_rows)
            ws.cell(row=row_idx, column=4, value=schema.total_size_display or "-")
            ws.cell(row=row_idx, column=5, value=round(schema.schema_quality_score, 4))

            grade = schema.schema_quality_grade
            cell = ws.cell(row=row_idx, column=6, value=grade)
            cell.fill = GRADE_FILLS.get(grade, GRADE_FILLS["N/A"])

            for col in range(1, 7):
                ws.cell(row=row_idx, column=col).border = THIN_BORDER

        self._auto_width(ws)

    def _write_table_profile(self, wb: Workbook, profile: DatabaseProfile) -> None:
        """Tablo Profil sheeti."""
        ws = wb.create_sheet("Tablo Profil")
        headers = [
            "Sema", "Tablo", "Aciklama", "Tip", "Satir Sayisi", "Tahmini", "Boyut",
            "Kolon Sayisi", "Sampling", "Sample %",
            "Kalite Skoru", "Kalite Notu", "Sure (sn)",
        ]
        self._apply_header(ws, headers)

        row_idx = 2
        for schema in profile.schemas:
            for table in schema.tables:
                ws.cell(row=row_idx, column=1, value=table.schema_name)
                ws.cell(row=row_idx, column=2, value=table.table_name)
                ws.cell(row=row_idx, column=3, value=table.table_description)
                ws.cell(row=row_idx, column=4, value=table.table_type)
                ws.cell(row=row_idx, column=5, value=table.row_count)
                ws.cell(row=row_idx, column=6, value="Evet" if table.row_count_estimated else "Hayir")
                ws.cell(row=row_idx, column=7, value=table.table_size_display or "-")
                ws.cell(row=row_idx, column=8, value=table.column_count)
                ws.cell(row=row_idx, column=9, value="Evet" if table.sampled else "Hayir")
                ws.cell(row=row_idx, column=10, value=table.sample_percent or "")
                ws.cell(row=row_idx, column=11, value=round(table.table_quality_score, 4))

                grade_cell = ws.cell(row=row_idx, column=12, value=table.table_quality_grade)
                grade_cell.fill = GRADE_FILLS.get(table.table_quality_grade, GRADE_FILLS["F"])

                ws.cell(row=row_idx, column=13, value=table.profile_duration_sec)

                for col in range(1, 14):
                    ws.cell(row=row_idx, column=col).border = THIN_BORDER
                row_idx += 1

        self._auto_width(ws)

    def _write_column_profile(self, wb: Workbook, profile: DatabaseProfile) -> None:
        """Kolon Profil sheeti (ana veri)."""
        ws = wb.create_sheet("Kolon Profil")
        headers = [
            "Sema", "Tablo", "Kolon", "Aciklama", "Sira", "Veri Tipi", "Max Uzunluk",
            "Nullable", "PK", "FK",
            "NULL Sayisi", "NULL Orani", "Distinct Sayisi", "Distinct Orani",
            "Min", "Max",
            "Ortalama", "Std Sapma", "P25", "P50", "P75",
            "Kalite Skoru", "Kalite Notu", "Kalite Bayraklari",
        ]
        self._apply_header(ws, headers)

        row_idx = 2
        for schema in profile.schemas:
            for table in schema.tables:
                for col in table.columns:
                    ws.cell(row=row_idx, column=1, value=table.schema_name)
                    ws.cell(row=row_idx, column=2, value=table.table_name)
                    ws.cell(row=row_idx, column=3, value=col.column_name)
                    ws.cell(row=row_idx, column=4, value=col.column_description)
                    ws.cell(row=row_idx, column=5, value=col.ordinal_position)
                    ws.cell(row=row_idx, column=6, value=col.data_type)
                    ws.cell(row=row_idx, column=7, value=col.max_length or "")
                    ws.cell(row=row_idx, column=8, value=col.is_nullable)
                    ws.cell(row=row_idx, column=9, value="PK" if col.is_primary_key else "")
                    ws.cell(row=row_idx, column=10, value="FK" if col.is_foreign_key else "")
                    ws.cell(row=row_idx, column=11, value=col.null_count)
                    ws.cell(row=row_idx, column=12, value=col.null_ratio)
                    ws.cell(row=row_idx, column=13, value=col.distinct_count)
                    ws.cell(row=row_idx, column=14, value=col.distinct_ratio)
                    ws.cell(row=row_idx, column=15, value=col.min_value or "")
                    ws.cell(row=row_idx, column=16, value=col.max_value or "")
                    ws.cell(row=row_idx, column=17, value=col.mean or "")
                    ws.cell(row=row_idx, column=18, value=col.stddev or "")

                    p25 = col.percentiles.get("p25", "") if col.percentiles else ""
                    p50 = col.percentiles.get("p50", "") if col.percentiles else ""
                    p75 = col.percentiles.get("p75", "") if col.percentiles else ""
                    ws.cell(row=row_idx, column=19, value=p25)
                    ws.cell(row=row_idx, column=20, value=p50)
                    ws.cell(row=row_idx, column=21, value=p75)

                    ws.cell(row=row_idx, column=22, value=round(col.quality_score, 4))
                    grade_cell = ws.cell(row=row_idx, column=23, value=col.quality_grade)
                    grade_cell.fill = GRADE_FILLS.get(col.quality_grade, GRADE_FILLS["F"])
                    ws.cell(row=row_idx, column=24, value=", ".join(col.quality_flags))

                    # PK/FK row renklendirme
                    if col.is_primary_key:
                        for c in range(1, 25):
                            ws.cell(row=row_idx, column=c).fill = PK_FILL
                    elif col.is_foreign_key:
                        for c in range(1, 25):
                            ws.cell(row=row_idx, column=c).fill = FK_FILL

                    for c in range(1, 25):
                        ws.cell(row=row_idx, column=c).border = THIN_BORDER

                    row_idx += 1

        self._auto_width(ws)

    def _write_top_values(self, wb: Workbook, profile: DatabaseProfile) -> None:
        """Top Degerler sheeti."""
        ws = wb.create_sheet("Top Degerler")
        headers = ["Sema", "Tablo", "Kolon", "Deger", "Frekans", "Yuzde"]
        self._apply_header(ws, headers)

        row_idx = 2
        for schema in profile.schemas:
            for table in schema.tables:
                for col in table.columns:
                    for tv in col.top_n_values:
                        ws.cell(row=row_idx, column=1, value=table.schema_name)
                        ws.cell(row=row_idx, column=2, value=table.table_name)
                        ws.cell(row=row_idx, column=3, value=col.column_name)
                        ws.cell(row=row_idx, column=4, value=str(tv.get("value", ""))[:200])
                        ws.cell(row=row_idx, column=5, value=tv.get("frequency", 0))
                        ws.cell(row=row_idx, column=6, value=tv.get("pct", 0))
                        for c in range(1, 7):
                            ws.cell(row=row_idx, column=c).border = THIN_BORDER
                        row_idx += 1

        self._auto_width(ws)

    def _write_pattern_analysis(self, wb: Workbook, profile: DatabaseProfile) -> None:
        """Pattern Analiz sheeti."""
        ws = wb.create_sheet("Pattern Analiz")
        headers = ["Sema", "Tablo", "Kolon", "Pattern", "Eslesme Orani", "Dominant"]
        self._apply_header(ws, headers)

        row_idx = 2
        for schema in profile.schemas:
            for table in schema.tables:
                for col in table.columns:
                    if not col.detected_patterns:
                        continue
                    for pattern_name, ratio in col.detected_patterns.items():
                        ws.cell(row=row_idx, column=1, value=table.schema_name)
                        ws.cell(row=row_idx, column=2, value=table.table_name)
                        ws.cell(row=row_idx, column=3, value=col.column_name)
                        ws.cell(row=row_idx, column=4, value=pattern_name)
                        ws.cell(row=row_idx, column=5, value=ratio)
                        ws.cell(
                            row=row_idx, column=6,
                            value="Evet" if pattern_name == col.dominant_pattern else "",
                        )
                        for c in range(1, 7):
                            ws.cell(row=row_idx, column=c).border = THIN_BORDER
                        row_idx += 1

        self._auto_width(ws)

    def _write_outlier_report(self, wb: Workbook, profile: DatabaseProfile) -> None:
        """Outlier Rapor sheeti."""
        ws = wb.create_sheet("Outlier Rapor")
        headers = [
            "Sema", "Tablo", "Kolon", "Q1", "Q3", "IQR",
            "Alt Sinir", "Ust Sinir", "Outlier Sayisi", "Outlier Orani",
        ]
        self._apply_header(ws, headers)

        row_idx = 2
        for schema in profile.schemas:
            for table in schema.tables:
                for col in table.columns:
                    if not col.outlier_bounds:
                        continue
                    b = col.outlier_bounds
                    ws.cell(row=row_idx, column=1, value=table.schema_name)
                    ws.cell(row=row_idx, column=2, value=table.table_name)
                    ws.cell(row=row_idx, column=3, value=col.column_name)
                    ws.cell(row=row_idx, column=4, value=b.get("q1"))
                    ws.cell(row=row_idx, column=5, value=b.get("q3"))
                    ws.cell(row=row_idx, column=6, value=b.get("iqr"))
                    ws.cell(row=row_idx, column=7, value=b.get("lower"))
                    ws.cell(row=row_idx, column=8, value=b.get("upper"))
                    ws.cell(row=row_idx, column=9, value=col.outlier_count)
                    ws.cell(row=row_idx, column=10, value=col.outlier_ratio)
                    for c in range(1, 11):
                        ws.cell(row=row_idx, column=c).border = THIN_BORDER
                    row_idx += 1

        self._auto_width(ws)
